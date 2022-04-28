
# Authors: Connor Castleberry, Max Wilson, Hannah Braslawsce, John Scheer
# Date: 4 / 26 / 2022
# Description: MA 305 Class Project, Covid statistics analyzer

# Note: COVID19PY code must be in the same folder as this program

# imports data
#import COVID19Py
import numpy as np
#import urllib3
import io
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from itertools import islice
from openpyxl import load_workbook
# imports user-defined functions
#import UserInput
import print_covid_graph as plotter

#vaccineType, dateRange, compareVaccine, vaccineType2, State = UserInput.UserInput()


###############################################################################
# #Test Vaccines report section
# # Note User Inputs section must be commented out inorder to use Test modes
# vaccineType = 'Pfizer'
# vaccineType2 = 'Moderna'
# State = None
# dateRange = None
# Dates = 0

# #Test mode two Graph
# dateRange = ['6/10/2021', '11/3/2021']
# vaccineType = None
# vaccineType2 = None
# State = None

#Test mode 3
# vaccineType = 'Moderna'
###############################################################################
###############################################################################
# User Inputs
optionPick = 6

while optionPick != '0':
    optionPick = input("Would you like to review total of people vaccinated for a vaccine(s) (1), hospitalizations per state (2), "
                       "or graph of new cases of COVID-19 by date? (3)\nTo quit program enter (0):\n")
    if optionPick == '1':
        vaccine1 = input("Please enter the name of the vaccine you would like to review (Moderna, J&J or Pfizer): ")
        print("Vaccine Type: " + vaccine1)
        if (vaccine1 == "Moderna") or (vaccine1 == "Pfizer") or (vaccine1 == "J&J"):
            print("Would you like to compare another vaccine against the one you have chosen? ")
            compareVaccine = input("Please enter yes or no: ")
            print("Answer: ", compareVaccine)
            global vaccineType
            vaccineType = vaccine1
            global State
            State = None
            if compareVaccine == "Yes" or compareVaccine == "yes":
                vaccine2 = input("Please enter the other vaccine you'd like to compare: ")
                print("You will be comparing " + vaccine1 + ' and ' + vaccine2)
                
                if (vaccine2 == "Moderna") or (vaccine2 == "Pfizer") or (vaccine2 == "J&J"):
                    global vaccineType2 
                    vaccineType2 = vaccine2
                    optionPick = '0'
            else:
                vaccineType2 = None
                optionPick = '0'
        else:
            print("Please enter a valid vaccine name.")
    elif optionPick == '2':
        State = input("Please enter which state you would like to review: ")
        print("State chosen: ", State)
        #hstateNums(hospitalStateNumbers)
        vaccineType = None
        vaccineType2 = None
        optionPick = '0'
    elif optionPick == '3':
        global dateRange
        Uses_dates = input("Would you like to see the graph between specific dates?\nyes or no:")
        if Uses_dates == 'yes' or (Uses_dates == 'Yes'):
            dateRange = input("\nPlease enter a date: ")
            print("Date Range: " + dateRange)
            print("The number of new-cases per day is shown below")
            print("See the graph below for a more detailed breakdown by day.")
            vaccineType = None
            vaccineType2 = None
            State = None
            optionPick = '0'
        else:
            print("See the graph below for a more detailed breakdown by day.")
            vaccineType = None
            vaccineType2 = None
            State = None
            optionPick = '0'
#end user inputs
###############################################################################       
        
# new sheet grab
# get data and graph

if bool(vaccineType) and not bool(vaccineType2):   #if vaccine one is requested
    wb2 = load_workbook(filename = 'covid19_vaccinations_in_the_united_states (1).xlsx')
    sheet_cases3 = wb2['Table1']

    data = sheet_cases3.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    full_excel_data = pd.DataFrame(data, index=idx, columns=cols)
    V1_total = full_excel_data.loc['United States', vaccineType]
    #print Vaccine 1 info
    Title = ' '.join(['The total number of people Vaccinated with', vaccineType,'is:', str(V1_total)])
    print(Title)
elif bool(State): #state data requested
    #Title = ' '.join(['Hospitalizations by state between', str(Dates[0]), 'and', str(Dates[-1])])
    print('Title')
elif bool(vaccineType2):  #if only one vaccine requested
    wb2 = load_workbook(filename = 'covid19_vaccinations_in_the_united_states (1).xlsx')
    sheet_cases3 = wb2['Table1']

    data = sheet_cases3.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    full_excel_data = pd.DataFrame(data, index=idx, columns=cols)
    V1_total = full_excel_data.loc['United States', vaccineType]
    V2_total = full_excel_data.loc['United States', vaccineType2]
    #
    #print info
    Title = ' '.join(['The total number of people Vaccinated with', vaccineType,'is:', str(V1_total)])
    print(Title)
    Title = ' '.join(['The total number of people Vaccinated with', vaccineType2,'is:', str(V2_total)])
    print(Title)
else:   # Cases Data requested
    wb2 = load_workbook(filename = 'data_table_for_daily_case_trends__the_united_states.xlsx')
    sheet_cases2 = wb2['Sheet2']

    data = sheet_cases2.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    full_excel_data = pd.DataFrame(data, index=idx, columns=cols)
    Dates_array = full_excel_data["Dates"].to_numpy()
    cases_data = full_excel_data["New Cases"]
    
    # clean up dates data
    count = 0
    Dates2 = []
    Dates = []
    for i in Dates_array:
        Dates2.append(str(i))
    for i in Dates2:
        string = ''.join([i[5:7], '/', i[8:10], '/', i[0:4]])
        if string != '//NaT': # make sure all data in the write format
            Dates.append(string)
    if bool(dateRange):  # no date range given
        #title and number evenout.
        Numbers = cases_data[0:-1]
        Title = ' '.join(['Covid Case between', Dates[0], 'and', Dates[-1]])
        #Graph
        plotter.print_covid_graph(Dates, Numbers, Title)
        
    else:   #date range given
        # collect right date range
        Dates_range = []
        first_date, last_date = dateRange
        for i in range(len(np.array(Dates ,dtype=object))):
            if int(first_date[-1]) == int(Dates[i][-1]): #if start and end dates are in same year
                if int(first_date[0]) <= int(Dates[i][0]):  #if first digit to digits are the same
                    if first_date[1] != '/': #if month is one digit
                        if first_date[3] != '/': #if day digit is one digit
                            if int(first_date[2]) <= int(Dates[i][0]): 
                                Dates_range.append(Dates[i])
                    else: #if month is two digits
                        if first_date[4] != '/': #if day is one digits
                            if int(first_date[3]) >= int(Dates[i][4]) or int(first_date[3]) >= int(Dates[i][4:5]):
                                Dates_range.append(Dates[i])
            elif int(first_date[-1]) <= int(Dates[i][-1]) and int(last_date[-1]) >= int(Dates[i][-1]): #if start and end dates are in different years
                if int(first_date[0]) <= int(Dates[i][0]):  #if month is big in Dates
                    if first_date[1] != '/': #if month is one digit
                        if first_date[3] != '/': #if day digit is one digit
                            if int(first_date[2]) <= int(Dates[i][0]): 
                                Dates_range.append(Dates[i])
                    else: #if month is two digits
                        if first_date[4] != '/': #if day is one digits
                            if int(first_date[3]) >= int(Dates[i][4]) or int(first_date[3]) >= int(Dates[i][4:5]):
                                Dates_range.append(Dates[i])
        
        #title and number evenout.
        Numbers = cases_data[0:-1]
        Title = ' '.join(['Covid Case between', Dates_range[0], 'and', Dates_range[-1]])
        
        #Graph
        plotter.print_covid_graph(Dates, Numbers, Title)

###########################################################################
